import os
import sys
import ChatBotModel
import openpyxl
import Card
from telegram.ext import Filters

def funcname():
    return sys._getframe(1).f_code.co_name

def callername():
    return sys._getframe(2).f_code.co_name

def proc_exit(bot, update):
    alpaca.reset()

def proc_stop(bot, update):
    alpaca.stop()

def proc_start(bot, update):
    alpaca.start()

def str_trimer(message):
    re = message.replace(" ","_")
    re = re.replace("\t","_")
    re = re.replace("__","_")
    re = re.replace("__","_")
    re = re.replace("__","_")
    re = re.replace("__","_")
    if '_' in re[0]:
        re = re[1:]
    return re

# document 수신시 발생 하는 이벤트
def get_document(bot, update) :
    if alpaca.file_caht:
        document_id = update.message.document.file_id
        document_file = bot.getFile(document_id)
        name = document_file.download("")
        r_file_name = make_report_sw(name)
        os.remove(name)
        if r_file_name:
            alpaca.sendDocument(document=open(r_file_name, 'rb'))
            os.remove(r_file_name)

def get_text(bot, update):
    if alpaca.file_caht:
        switch_text(update.message.text)

def switch_text(text):
	print(Card.HotCommand(text))
    if 'proc_hyundea_card' in alpaca.event_type:
        return make_report_hyundea_card(text)

def proc_hyundea_card(bot, update):
    alpaca.update_index(funcname())

def make_report_hyundea_card(text):
    item = text.split('\n')

    b_start = False
    b_cash = False
    b_world = False
    b_write = False
    output = ""
    store = ""
    date = ""
    value = ""
    count = 0

    for line in item:
        #[Web발신]\n[현대카드] 해외승인\n주*훈님\n12/20 09:51\nKRW 13,500.00\nITUNES.COM/BILL\n*원화결제\n
        #[Web발신]\n현대카드 the Green 승인\n주*훈\n29,200원 일시불\n12/14 13:24\n다이소아성산\n누적847,780원\n
        if '[Web발신]' in line:
            b_start = True
            b_cash = False
            b_world = False
            b_write = False
            store = ""
            date = ""
            value = ""
            count = 0
        elif b_start:
            count = count + 1
            if "현대카드 the Green 승인" in line and count == 1:
                b_cash = True
            elif "[현대카드] 해외승인" in line and count == 1:
                b_world = True
            elif b_cash:
                if "일시불" in line and count == 3:
                    value = line.replace("원 일시불", "")
                    value = value.replace(",", "")
                elif not date and count == 4:
                    date = line
                elif not store and count == 5:
                    store = line
            elif b_world:
                if "KRW" in line and count == 4:
                    value = line.replace("KRW ", "")
                    value = value.replace(",", "")
                elif not date and count == 3:
                    date = line
                elif not store and count == 5:
                    store = line
        
        if date and store and value and not b_write:
            b_write = True
            output = output + store+", " + value + ", " + date + "\n"

    alpaca.sendMessage(output)

def make_report_sw(name):
    if 'proc_signal_log_ci' in alpaca.event_type:
        return make_report_signal_log_ci(name)
    elif 'proc_signal_log_cus' in alpaca.event_type:
        return make_report_signal_log_cus(name)
    elif 'proc_infomation_log' in alpaca.event_type:
        return make_report_infomation_log(name)

def proc_signal_log_ci(bot, update):
    alpaca.update_index(funcname())

def make_report_signal_log_ci(name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['ModulID','Signal Category','No','Log Type','Signal Name','Description', 'ifdef..endif'])
    case_name = []

    with open(name, mode='rt', encoding='euc-kr') as f:
        offset = 9
        for line in f:
            start = line.find('const int')
            end = line.find('=')
            dummy = 'DUMMY' in line
            is_able = start is not -1 and end is not -1 and not dummy
            if is_able:
                signal_name = line[start+offset:end].strip()
                if 'IN_' in signal_name:
                    signal_name = signal_name.replace("IN_", "X_")
                    signal_categroy = 'Input'
                else:
                    signal_name = signal_name.replace("OUT_", "Y_")
                    signal_categroy = 'Output'
                signal_name = signal_name.replace("__", "_")
                log_type = "S"
                number = ''
                module_id = ''
                description = signal_name[2:]
                ws.append([module_id,signal_categroy,number,log_type,signal_name, description] + case_name)
            elif '#' in line:
                if 'endif' in line:
                    del case_name[-1]
                elif 'else' in line:
                    case_name[-1] = case_name[-1].replace('elif_','else_')
                    case_name[-1] = case_name[-1].replace('if_','else_')
                elif 'elif' in line:
                    case_name[-1] = line[line.find('elif')+4:].strip()
                    case_name[-1] = case_name[-1].replace('\t','')
                    case_name[-1] = 'elif_' + case_name[-1]
                elif 'if' in line:
                    case_name.append(line[line.find('if')+2:].strip())
                    case_name[-1] = case_name[-1].replace('\t','')
                    case_name[-1] = 'if_' + case_name[-1]

    # 엑셀 파일 저장
    result_file_name = "score2.xlsx"
    wb.save(result_file_name)
    wb.close()
    return result_file_name

def proc_signal_log_cus(bot, update):
    alpaca.update_index(funcname())

def make_report_signal_log_cus(name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['ModulID','Signal Category','No','Log Type','Signal Name','Description', 'ifdef..endif'])
    case_name = []

    with open(name, mode='rt', encoding='euc-kr') as f:
        offset = 20
        for line in f:
            start = line.find('const unsigned short')
            end = line.find('=')
            dummy = 'DUMMY' in line
            is_able = start is not -1 and end is not -1 and not dummy
            if is_able:
                signal_name = line[start+offset:end].strip()
                if 'IN_' in signal_name:
                    signal_name = signal_name.replace("IN_", "X_")
                    signal_categroy = 'Input'
                else:
                    signal_name = signal_name.replace("OUT_", "Y_")
                    signal_categroy = 'Output'
                signal_name = signal_name.replace("__", "_")
                log_type = "S"
                number = ''
                module_id = ''
                description = signal_name[2:]
                ws.append([module_id,signal_categroy,number,log_type,signal_name, description] + case_name)
            elif '#' in line:
                if 'endif' in line:
                    del case_name[-1]
                elif 'else' in line:
                    case_name[-1] = case_name[-1].replace('elif_','else_')
                    case_name[-1] = case_name[-1].replace('if_','else_')
                elif 'elif' in line:
                    case_name[-1] = line[line.find('elif')+4:].strip()
                    case_name[-1] = case_name[-1].replace('\t','')
                    case_name[-1] = 'elif_' + case_name[-1]
                elif 'if' in line:
                    case_name.append(line[line.find('if')+2:].strip())
                    case_name[-1] = case_name[-1].replace('\t','')
                    case_name[-1] = 'if_' + case_name[-1]

    # 엑셀 파일 저장
    result_file_name = "score2.xlsx"
    wb.save(result_file_name)
    wb.close()
    return result_file_name

def proc_infomation_log(bot, update):
    alpaca.update_index(funcname())

def make_report_infomation_log(name):
    
    units = []
    positions = []

    with open(name, mode='rt', encoding='euc-kr') as f:
        for line in f:
            dummy = ';' in line
            is_able = not dummy
            if is_able:
                l_data = line.split(',')

                if 'UNIT' in l_data[0]:
                    units.append(l_data)
                elif 'POS' in l_data[0]:
                    positions.append(l_data)


    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['ModulID','Information Category','No','Log Type (I)','Parameter Name','Description', 'From Position List(ModuleID#SlotNo)', 'To Position List(ModuleID#SlotNo)', 'unit', 'Min', 'Max'])

    temp = []
    temp.append("") #module id
    temp.append("ALL") #information category
    temp.append("") #no
    temp.append("I") #log type
    temp.append("") #parameter name
    temp.append("") #description
    temp.append("") #from
    temp.append("") #to
    temp.append("") #unit
    temp.append("") #min
    temp.append("") #max

    for line in units:
        for i in range(int(line[3].strip())):
            temp[4] = str_trimer(line[2].upper() + '_' + line[4+i].upper() + "_TORQUE")
            temp[5] = line[2].lower() + " 유닛 " + line[4+i].lower() + " 축 부하율"
            temp[8] = "%"
            temp[9] = "0"
            temp[10] = "100"
            ws.append(temp)

            temp[4] = str_trimer(line[2].upper() + '_' + line[4+i].upper() + "_SPEED")
            temp[5] = line[2].lower() + " 유닛 " + line[4+i].lower() + " 축 현제 속도"
            temp[8] = "mm/s"
            temp[9] = ""
            temp[10] = ""
            ws.append(temp)

            temp[4] = str_trimer("SET_" + line[2].upper() + '_' + line[4+i].upper() + "_SPEED")
            temp[5] = line[2].lower() + " 유닛 " + line[4+i].lower() + " 축 설정 속도"
            temp[8] = "mm/s"
            temp[9] = ""
            temp[10] = ""
            ws.append(temp)

            temp[4] = str_trimer(line[2].upper() + '_' + line[4+i].upper() + "_CUR_POS")
            temp[5] = line[2].lower() + " 유닛 " + line[4+i].lower() + " 축 현제 위치"
            temp[8] = "mm"
            temp[9] = ""
            temp[10] = ""
            ws.append(temp)

            for pos in positions:
                if pos[1].strip() == line[1].strip():
                    temp[4] = str_trimer("SET_" + line[2].upper() + '_' + line[4+i].upper() + '_' + pos[5].upper() + "_POS")
                    temp[5] = line[2].lower() + " 유닛 " + line[4+i].lower() + " 축 " + pos[5].lower() + " 설정 위치"
                    temp[8] = "mm"
                    temp[9] = ""
                    temp[10] = ""
                    ws.append(temp)

    result_file_name = "score2.xlsx"
    wb.save(result_file_name)
    wb.close()
    return result_file_name


alpaca = ChatBotModel.Bot()

alpaca.add_command_handler('stop', proc_stop)
alpaca.add_command_handler('start', proc_start)
alpaca.add_command_handler('exit', proc_exit)
alpaca.add_command_handler_level('proc_signal_log_ci', proc_signal_log_ci)
alpaca.add_command_handler_level('proc_signal_log_cus', proc_signal_log_cus)
alpaca.add_command_handler_level('proc_infomation_log', proc_infomation_log)
alpaca.add_command_handler_level('proc_hyundea_card', proc_hyundea_card)
alpaca.add_message_handler(Filters.document, get_document)
alpaca.add_message_handler(Filters.text, get_text)

alpaca.start()